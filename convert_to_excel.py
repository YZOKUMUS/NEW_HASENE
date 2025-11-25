#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TEST_PLANI.md dosyasını Excel (.xlsx) formatına dönüştürür
"""

import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean_text(text):
    """Markdown formatından temizleme"""
    if not text:
        return ""
    # Emoji'leri kaldır (başta veya sonda)
    emoji_pattern = r'[📚🎧📝📖🤲📜💰⭐🏆🥇📊📅🎯🔥🏅📈🎨📱🎭🔊💾🔄🗑️📤📥📲📴🔄🛡️🔒🚫⚡🎮💾🌐🦊🍎🪟📱🤖✅❌🔴🟡🟢⚪📋🎯]'
    text = re.sub(emoji_pattern, '', text)
    text = re.sub(r'^\d+\.\s*', '', text, flags=re.MULTILINE)  # Numaralandırmaları kaldır
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)  # **bold** -> bold
    text = re.sub(r'`([^`]+)`', r'\1', text)  # `code` -> code
    text = re.sub(r'```[\s\S]*?```', '', text)  # Code bloklarını kaldır
    text = re.sub(r'^#{1,6}\s+', '', text, flags=re.MULTILINE)  # Başlıkları temizle
    text = re.sub(r'^---+\s*$', '', text, flags=re.MULTILINE)  # Ayırıcıları kaldır
    # Liste işaretlerini kaldır ama içeriği koru
    text = re.sub(r'^\s*[-*]\s+', '', text, flags=re.MULTILINE)  # Liste işaretlerini kaldır
    text = re.sub(r'\[([^\]]+)\]\([^\)]+\)', r'\1', text)  # Linkleri temizle
    # ✅ işaretini kaldır
    text = re.sub(r'✅\s*', '', text)
    text = text.strip()
    return text

def parse_markdown(md_content):
    """Markdown içeriğini parse et ve testleri çıkar"""
    tests = []
    current_category = ""
    current_test = None
    
    lines = md_content.split('\n')
    i = 0
    
    while i < len(lines):
        line = lines[i].strip()
        
        # Kategori başlıkları (## ile başlayan)
        if line.startswith('##') and not line.startswith('###'):
            current_category = clean_text(line)
            i += 1
            continue
        
        # Test başlıkları (#### ile başlayan veya ### ile başlayan ve Test içeren)
        if (line.startswith('####') or (line.startswith('###') and 'Test' in line)):
            # Önceki testi kaydet
            if current_test:
                tests.append(current_test)
            
            # Yeni test başlat
            test_name = clean_text(line)
            test_number = ""
            # Test numarasını bul (Test 1.1, Test 2.3 gibi)
            match = re.search(r'Test\s+(\d+\.\d+)', test_name)
            if match:
                test_number = match.group(1)
                # Test numarasını ve iki nokta üst üsteyi kaldır
                test_name = re.sub(r'Test\s+\d+\.\d+\s*:?\s*', '', test_name, count=1).strip()
                # Emoji'leri kaldır
                test_name = re.sub(r'^[📚🎧📝📖🤲📜💰⭐🏆🥇📊📅🎯🔥🏅📈🎨📱🎭🔊💾🔄🗑️📤📥📲📴🔄🛡️🔒🚫⚡🎮💾🌐🦊🍎🪟📱🤖]\s*', '', test_name)
            
            current_test = {
                'number': test_number,
                'name': test_name,
                'category': current_category,
                'steps': "",
                'expected': "",
                'checkpoints': ""
            }
            i += 1
            continue
        
        # Test içeriği
        if current_test:
            if line.startswith('**Ne Yapayım:**'):
                i += 1
                steps = []
                while i < len(lines) and not lines[i].strip().startswith('**Ne Kazanırım:**'):
                    step_line = lines[i].strip()
                    if step_line and not step_line.startswith('**'):
                        steps.append(clean_text(step_line))
                    i += 1
                current_test['steps'] = '\n'.join(steps)
                continue
            
            elif line.startswith('**Ne Kazanırım:**'):
                i += 1
                expected = []
                while i < len(lines) and not lines[i].strip().startswith('**Nereyi Kontrol Edip Ne Görmeliyim:**'):
                    exp_line = lines[i].strip()
                    if exp_line and not exp_line.startswith('**'):
                        expected.append(clean_text(exp_line))
                    i += 1
                current_test['expected'] = '\n'.join(expected)
                continue
            
            elif line.startswith('**Nereyi Kontrol Edip Ne Görmeliyim:**'):
                i += 1
                checkpoints = []
                while i < len(lines):
                    cp_line = lines[i].strip()
                    if cp_line.startswith('---') or (cp_line.startswith('##') and not cp_line.startswith('###')):
                        break
                    if cp_line.startswith('####') or (cp_line.startswith('###') and 'Test' in cp_line):
                        break
                    if cp_line and not cp_line.startswith('**'):
                        checkpoints.append(clean_text(cp_line))
                    i += 1
                current_test['checkpoints'] = '\n'.join(checkpoints)
                continue
        
        i += 1
    
    # Son testi ekle
    if current_test:
        tests.append(current_test)
    
    return tests

def create_excel(tests, output_file='TEST_PLANI.xlsx'):
    """Excel dosyası oluştur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Planı"
    
    # Başlık satırı
    headers = ['Test No', 'Kategori', 'Test Adı', 'Ne Yapayım', 'Ne Kazanırım', 
               'Nereyi Kontrol Edip Ne Görmeliyim', 'Durum', 'Notlar', 'Tarih', 'Test Eden']
    
    # Stil tanımlamaları
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Başlıkları yaz
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Test verilerini yaz
    for row_num, test in enumerate(tests, 2):
        ws.cell(row=row_num, column=1, value=test.get('number', '')).border = border
        ws.cell(row=row_num, column=2, value=test.get('category', '')).border = border
        ws.cell(row=row_num, column=3, value=test.get('name', '')).border = border
        ws.cell(row=row_num, column=4, value=test.get('steps', '')).border = border
        ws.cell(row=row_num, column=5, value=test.get('expected', '')).border = border
        ws.cell(row=row_num, column=6, value=test.get('checkpoints', '')).border = border
        ws.cell(row=row_num, column=7, value='').border = border  # Durum
        ws.cell(row=row_num, column=8, value='').border = border  # Notlar
        ws.cell(row=row_num, column=9, value='').border = border  # Tarih
        ws.cell(row=row_num, column=10, value='').border = border  # Test Eden
        
        # Hücre stilleri
        for col_num in range(1, 11):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col_num in [4, 5, 6, 8]:  # Uzun metin sütunları
                cell.alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')
    
    # Sütun genişliklerini ayarla
    ws.column_dimensions['A'].width = 10  # Test No
    ws.column_dimensions['B'].width = 25  # Kategori
    ws.column_dimensions['C'].width = 30  # Test Adı
    ws.column_dimensions['D'].width = 40  # Ne Yapayım
    ws.column_dimensions['E'].width = 35  # Ne Kazanırım
    ws.column_dimensions['F'].width = 50  # Kontrol Noktaları
    ws.column_dimensions['G'].width = 12  # Durum
    ws.column_dimensions['H'].width = 30  # Notlar
    ws.column_dimensions['I'].width = 12  # Tarih
    ws.column_dimensions['J'].width = 15  # Test Eden
    
    # Satır yüksekliklerini ayarla
    ws.row_dimensions[1].height = 30
    for row_num in range(2, len(tests) + 2):
        ws.row_dimensions[row_num].height = 60
    
    # Filtre ekle
    ws.auto_filter.ref = ws.dimensions
    
    # Kaydet
    wb.save(output_file)
    print(f"✅ Excel dosyası oluşturuldu: {output_file}")
    print(f"📊 Toplam {len(tests)} test eklendi")

if __name__ == '__main__':
    try:
        # Markdown dosyasını oku
        with open('TEST_PLANI.md', 'r', encoding='utf-8') as f:
            md_content = f.read()
        
        # Parse et
        print("📖 Markdown dosyası okunuyor...")
        tests = parse_markdown(md_content)
        print(f"✅ {len(tests)} test bulundu")
        
        # Excel oluştur
        print("📊 Excel dosyası oluşturuluyor...")
        create_excel(tests, 'TEST_PLANI.xlsx')
        
    except FileNotFoundError:
        print("❌ Hata: TEST_PLANI.md dosyası bulunamadı!")
    except Exception as e:
        print(f"❌ Hata: {str(e)}")
        import traceback
        traceback.print_exc()

